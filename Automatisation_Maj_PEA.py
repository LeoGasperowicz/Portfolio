import pandas as pd
import yfinance as yf
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font
import schedule
import time

def process_excel_file():
    file_path = r"C:\Users\Utilisateur\Desktop\Optimization PEA\PEA_Fr.xlsx"  # Remplacez par le chemin vers votre fichier Excel
    df = pd.read_excel(file_path)
    print(df.columns)
    print(df.head(10))

    # Liste des symboles boursiers des actions
    symbols = ['TFI.PA', 'TTE.PA', 'BB.PA', 'ML.PA', 'SGO.PA', 'STLAP.PA']

    # Récupérer les prix de clôture les plus récents disponibles
    current_prices = yf.download(symbols, period="5d")['Adj Close'].iloc[-1]

    # Obtenir la date actuelle
    current_date = datetime.now().strftime("%Y-%m-%d")

    # Correspondance entre les noms des actions dans le DataFrame et les symboles téléchargés
    name_to_symbol = {
        'TF1 (TF1)': 'TFI.PA',
        'Total Énergie (TTE)': 'TTE.PA',
        'BIC (BIF)': 'BB.PA',
        'Michelin (ML)': 'ML.PA',
        'Saint Gobain (SGO)': 'SGO.PA',
        'STELLANTIS (STLAP)': 'STLAP.PA'
    }

    # Créer de nouvelles colonnes avec des noms uniques basés sur la date
    update_price_col = f'Update Prix {current_date}'
    montant_actuel_col = f'Montant actuel {current_date}'
    pourcentage_augmentation_col = f'Pourcentage d\'augmentation {current_date}'

    df[update_price_col] = df['Nom de l\'action'].map(name_to_symbol).map(current_prices)
    df[montant_actuel_col] = df[update_price_col] * df['Quantity']
    df[pourcentage_augmentation_col] = ((df[montant_actuel_col] / df['Montant acheté']) - 1) * 100

    # Enregistrer les modifications dans le DataFrame
    df.to_excel(file_path, index=False)

    # Charger le fichier Excel pour appliquer les styles
    wb = load_workbook(file_path)
    ws = wb.active

    # Définir les styles
    gris_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    vert_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    rouge_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    jaune_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    orange_fill = PatternFill(start_color='FFCC99', end_color='FFCC99', fill_type='solid')

    border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
    header_font = Font(bold=True)

    # Appliquer les bordures et le formatage des en-têtes
    for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=ws.max_column):
        for cell in col:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border_style

    # Appliquer le style "vérification" à la colonne 'Montant acheté'
    montant_achete_col_idx = df.columns.get_loc('Montant acheté') + 1
    for cell in ws.iter_cols(min_col=montant_achete_col_idx, max_col=montant_achete_col_idx, min_row=2, max_row=7):
        for c in cell:
            c.fill = gris_fill
            c.border = border_style

    # Appliquer les styles conditionnels pour les pourcentages d'augmentation
    augmentation_cols = [col for col in df.columns if 'Pourcentage d\'augmentation' in col]
    for col in augmentation_cols:
        col_idx = df.columns.get_loc(col) + 1
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, max_row=7):
            for c in cell:
                if c.value is not None:
                    if float(c.value) > 0:
                        c.fill = vert_fill
                    else:
                        c.fill = rouge_fill
                c.border = border_style

    # Afficher les noms des colonnes pour le débogage
    print("Colonnes dans le DataFrame après ajout:", df.columns)

    # Trouver la ligne de total
    total_row_idx = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            if cell.value == 'Total':
                total_row_idx = cell.row
                break
        if total_row_idx:
            break

    # Calculer et appliquer les styles aux totaux
    # Écrire les totaux pour la colonne 'Montant acheté'
    ws.cell(row=total_row_idx, column=montant_achete_col_idx).value = sum([ws.cell(row=i, column=montant_achete_col_idx).value for i in range(2, 8)])
    ws.cell(row=total_row_idx, column=montant_achete_col_idx).fill = gris_fill
    ws.cell(row=total_row_idx, column=montant_achete_col_idx).border = border_style

    # Calculer et écrire les totaux pour chaque colonne de 'Montant actuel' et 'Pourcentage d'augmentation'
    for col in augmentation_cols:
        montant_actuel_col = col.replace('Pourcentage d\'augmentation', 'Montant actuel')
        print(f"Traitement de la colonne: {col}, correspondant à la colonne: {montant_actuel_col}")  # Ligne de débogage
        if montant_actuel_col not in df.columns:
            print(f"Colonne {montant_actuel_col} non trouvée dans le DataFrame")  # Ligne de débogage
            continue  # Passer à la prochaine itération si la colonne n'existe pas

        montant_actuel_col_idx = df.columns.get_loc(montant_actuel_col) + 1
        montant_actuel_total = sum([ws.cell(row=i, column=montant_actuel_col_idx).value for i in range(2, 8)])
        ws.cell(row=total_row_idx, column=montant_actuel_col_idx).value = montant_actuel_total
        ws.cell(row=total_row_idx, column=montant_actuel_col_idx).border = border_style

        # Obtenir la dernière colonne de 'Montant actuel'
        last_montant_actuel_col = [col for col in df.columns if 'Montant actuel' in col][-2]
        last_montant_actuel_total = sum([ws.cell(row=i, column=df.columns.get_loc(last_montant_actuel_col) + 1).value for i in range(2, 8)])

        # Appliquer les styles conditionnels pour les totaux
        if montant_actuel_total > last_montant_actuel_total and montant_actuel_total > sum([ws.cell(row=i, column=montant_achete_col_idx).value for i in range(2, 8)]):
            ws.cell(row=total_row_idx, column=montant_actuel_col_idx).fill = vert_fill
        elif montant_actuel_total < last_montant_actuel_total and montant_actuel_total > sum([ws.cell(row=i, column=montant_achete_col_idx).value for i in range(2, 8)]):
            ws.cell(row=total_row_idx, column=montant_actuel_col_idx).fill = jaune_fill
        elif montant_actuel_total > last_montant_actuel_total and montant_actuel_total < sum([ws.cell(row=i, column=montant_achete_col_idx).value for i in range(2, 8)]):
            ws.cell(row=total_row_idx, column=montant_actuel_col_idx).fill = orange_fill
        else:
            ws.cell(row=total_row_idx, column=montant_actuel_col_idx).fill = rouge_fill

        # Calculer et appliquer les styles conditionnels pour les pourcentages d'augmentation
        pourcentage_augmentation_col = col
        pourcentage_augmentation_col_idx = df.columns.get_loc(pourcentage_augmentation_col) + 1
        
        # Corrected calculation for total percentage increase
        pourcentage_augmentation_total = ((montant_actuel_total / last_montant_actuel_total) - 1) * 100
        
        ws.cell(row=total_row_idx, column=pourcentage_augmentation_col_idx).value = pourcentage_augmentation_total
        ws.cell(row=total_row_idx, column=pourcentage_augmentation_col_idx).border = border_style

        last_pourcentage_augmentation_col = [col for col in df.columns if 'Pourcentage d\'augmentation' in col][-2]
        last_pourcentage_augmentation_total = ws.cell(row=total_row_idx, column=df.columns.get_loc(last_pourcentage_augmentation_col) + 1).value

        if pourcentage_augmentation_total > last_pourcentage_augmentation_total:
            ws.cell(row=total_row_idx, column=pourcentage_augmentation_col_idx).fill = vert_fill
        elif pourcentage_augmentation_total < last_pourcentage_augmentation_total and pourcentage_augmentation_total > 0:
            ws.cell(row=total_row_idx, column=pourcentage_augmentation_col_idx).fill = jaune_fill
        elif pourcentage_augmentation_total > last_pourcentage_augmentation_total and pourcentage_augmentation_total < 0:
            ws.cell(row=total_row_idx, column=pourcentage_augmentation_col_idx).fill = orange_fill
        else:
            ws.cell(row=total_row_idx, column=pourcentage_augmentation_col_idx).fill = rouge_fill

    # Appliquer le format monétaire à toutes les colonnes contenant le mot 'prix' ou 'montant'
    for col in df.columns:
        if 'prix' in col.lower() or 'montant' in col.lower():
            col_idx = df.columns.get_loc(col) + 1
            for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, max_row=ws.max_row):
                for c in cell:
                    c.number_format = '#,##0.0000 €'

    # Ajuster la largeur des colonnes
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Enregistrer le fichier Excel
    wb.save(file_path)

# Configuration de la planification avec schedule
schedule.every().tuesday.at("17:16").do(process_excel_file)

print("La planification est configurée. Le script s'exécutera tous les Vendredi à 18h30.")

# Boucle infinie pour exécuter les tâches planifiées
while True:
    schedule.run_pending()
    time.sleep(1)
